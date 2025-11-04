#!/usr/bin/env python3
"""Generates comparison plots for the simple reflex and random agents."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import matplotlib

matplotlib.use("Agg")  # noqa: E402 keep backend config before pyplot import
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_FILE = BASE_DIR / "ejercicio3y4muestras.xlsx"
OUTPUT_DIR = BASE_DIR / "images"
AGENT_STYLES = {
    "Agente Aleatorio": {
        "color": "#d62728",
        "linestyle": "--",
        "marker": "s",
        "markersize": 6,
        "markerfacecolor": "none",
        "linewidth": 2.2,
        "zorder": 3,
    },
    "Agente reflexivo simple": {
        "color": "#1f77b4",
        "linestyle": "-",
        "marker": "o",
        "markersize": 6,
        "linewidth": 2.2,
        "zorder": 2,
    },
}


def _extract_records() -> pd.DataFrame:
    """Parse the Excel workbook into a tidy dataframe."""
    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Hoja 1"]

    records: List[Dict[str, float]] = []
    current_agent: str | None = None
    current_grid: str | None = None
    group_starts: List[int] = []

    for row in ws.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            continue

        first_cell = row[0]

        if isinstance(first_cell, str) and first_cell.startswith("Agente"):
            current_agent = first_cell.strip()
            current_grid = None
            group_starts = []
            continue

        if isinstance(first_cell, str) and "x" in first_cell:
            current_grid = first_cell.strip()
            group_starts = []

            idx = 1
            while idx < len(row):
                header = row[idx]
                if header is None:
                    idx += 1
                    continue
                group_starts.append(idx)
                idx += 6  # skip to the next block

            continue

        if not (current_agent and current_grid and group_starts):
            continue

        for start in group_starts:
            suciedad = row[start]
            if suciedad is None:
                continue

            records.append(
                {
                    "agent": current_agent,
                    "grid": current_grid,
                    "suciedad": suciedad,
                    "sucias": row[start + 1],
                    "limpio": row[start + 2],
                    "pct_limpieza": row[start + 3],
                    "acciones": row[start + 4],
                    "semilla": row[start + 5],
                }
            )

    return pd.DataFrame.from_records(records)


def _prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean column types and add helpers for plotting."""
    numeric_cols = ["suciedad", "sucias", "limpio", "pct_limpieza", "acciones", "semilla"]
    df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors="coerce")
    df = df.dropna(subset=["suciedad"])

    dims = df["grid"].str.split("x", expand=True).astype(int)
    df["grid_rows"] = dims[0]
    df["grid_cols"] = dims[1]
    df["grid_area"] = df["grid_rows"] * df["grid_cols"]

    df = df.sort_values(["grid_area", "suciedad", "agent", "semilla"]).reset_index(drop=True)
    return df


def _plot_metric(df: pd.DataFrame, metric: str, ylabel: str, title: str, output_path: Path) -> None:
    """Draw a grid of line plots comparing agents for a given metric."""
    grids = sorted(
        df["grid"].unique(),
        key=lambda g: (int(g.split("x")[0]), int(g.split("x")[1])),
    )

    fig, axes = plt.subplots(2, 3, figsize=(15, 8), sharex=True)
    axes = axes.flatten()

    handles = []
    labels = []
    agent_order = list(AGENT_STYLES.keys())

    for ax, grid in zip(axes, grids, strict=False):
        grid_data = df[df["grid"] == grid]
        grouped = (
            grid_data.groupby(["agent", "suciedad"], as_index=False)[metric]
            .mean()
            .sort_values("suciedad")
        )

        agent_groups = {agent: agent_data for agent, agent_data in grouped.groupby("agent")}

        ordered_agents = [
            agent for agent in agent_order if agent in agent_groups
        ] + [
            agent for agent in agent_groups.keys() if agent not in agent_order
        ]

        for agent in ordered_agents:
            agent_data = agent_groups[agent]
            style = AGENT_STYLES.get(agent, {})
            line, = ax.plot(
                agent_data["suciedad"],
                agent_data[metric],
                marker=style.get("marker", "o"),
                linestyle=style.get("linestyle", "-"),
                linewidth=style.get("linewidth", 2.0),
                markersize=style.get("markersize", 5),
                color=style.get("color"),
                markerfacecolor=style.get("markerfacecolor"),
                zorder=style.get("zorder", 2),
                label=agent,
            )
            if not handles:
                handles.append(line)
                labels.append(agent)
            elif agent not in labels:
                handles.append(line)
                labels.append(agent)

        ax.set_title(grid)
        ax.set_xlabel("Probabilidad de suciedad inicial")
        ax.set_ylabel(ylabel)
        ax.grid(True, alpha=0.3)

    for ax in axes[len(grids) :]:
        fig.delaxes(ax)

    if handles:
        fig.legend(handles, labels, loc="upper center", ncol=len(labels))

    fig.suptitle(title, fontsize=16)
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=300)
    plt.close(fig)


def main() -> None:
    df = _prepare_dataframe(_extract_records())

    _plot_metric(
        df,
        metric="pct_limpieza",
        ylabel="% de limpieza promedio",
        title="Porcentaje de limpieza promedio por agente y tamaño de entorno",
        output_path=OUTPUT_DIR / "comparacion_pct_limpieza.png",
    )

    _plot_metric(
        df,
        metric="acciones",
        ylabel="Acciones promedio",
        title="Acciones promedio por agente y tamaño de entorno",
        output_path=OUTPUT_DIR / "comparacion_acciones.png",
    )


if __name__ == "__main__":
    main()
